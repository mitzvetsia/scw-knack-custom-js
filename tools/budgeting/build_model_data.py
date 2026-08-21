#!/usr/bin/env python3
"""Stage 1: parse the Xero GL Detail export + legacy budgeting-tool exports into model_data.json.

Inputs (paths passed as args or edit DEFAULTS):
  1. Xero "General Ledger Detail" xlsx export (full year, accrual)
  2. contractsexpenses.csv  (legacy budget tool: Contracts & Expenses table)
  3. allocations.csv        (legacy budget tool: Allocations table)
  4. legacy BUDGET_TEST workbook (STAFFING / PROJECTED REVENUE CALCULATIONS / RAMP SCHEDULES)

Rules encoded here:
  - expenses are read from the P&L account sections (accrual: bills at bill date)
  - payroll accounts (66000/6529/6727/6189), COGS (5xxx + 6734) and revenue (4xxx) are
    aggregated separately from vendor-level opex
  - vendor names are canonicalized (payment journals, dated bill names, card-fee sources)
  - department: budgeting-tool allocation wins over Xero's Departments tracking category
  - per-seat (position-based) allocations are carried as seat_monthly -> modeled on STAFFING
"""
import sys, json, csv, re, collections, statistics, os
from datetime import datetime, date
import openpyxl

DEFAULTS = dict(
    gl="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/7cf4d3fa-Security_Camera_Warehouse__INC__General_Ledger_Detail_2.xlsx",
    contracts="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/ccd2bd72-contractsexpenses.csv",
    allocations="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/47bec46b-allocations.csv",
    legacy="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/25b0d1bd-BUDGET_TEST_1774983468__Q2Q3_2026_JULY_REWORK.xlsx",
    cutoff="2026-08-21",
)
A = dict(DEFAULTS)
for arg in sys.argv[1:]:
    k, _, v = arg.partition("=")
    A[k] = v
CUTOFF = A["cutoff"]
cut = date.fromisoformat(CUTOFF)
ELAPSED = (cut.month - 1) + cut.day / 31.0

def money(s):
    if s is None: return 0.0
    s = str(s).replace("$", "").replace(",", "").strip()
    if not s: return 0.0
    neg = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = s.strip("()-")
    try: v = float(s)
    except ValueError: return 0.0
    return -v if neg else v

# ---------------- 1. parse GL ----------------
wb = openpyxl.load_workbook(A["gl"], read_only=True, data_only=True)
ws = wb.worksheets[0]
records, section = [], None
SKIP = ("Total","Opening","Closing","Net movement","No transactions","Date","General Ledger","For the period")
for row in ws.iter_rows(values_only=True):
    a = row[0]
    rest = [v for v in row[1:] if v not in (None, "")]
    if isinstance(a, str) and not rest and not a.startswith(SKIP) and "period" not in a:
        if not re.match(r"^[A-Za-z ]+,? (INC|LLC|Inc)", a):
            section = a.strip()
        continue
    if isinstance(a, datetime) and a.date().isoformat() <= CUTOFF:
        _, source, desc, ref, debit, credit, runbal, dept, proj, relacct = row[:10]
        records.append((section, a.month, (desc or "").strip(), float(debit or 0), float(credit or 0), (dept or "").strip()))
wb.close()

def code(acct): return acct.split(" - ")[0].strip() if acct else ""
PAYROLL_PL = {"66000","6529","6727","6189"}
COGS = {"5000","5257","5258","5259","5260","5566","5567","5568","5569","5570","5571","6734"}
REVENUE_SALES = {"4048","4049","4052","4053","4054","4056","4057","4058","4059","4060","4062"}
OTHER_INCOME = {"4050","4051","8311"}
def is_opex(c):
    return (c.startswith(("6","8")) or c == "208") and c not in PAYROLL_PL and c not in COGS and c != "8311"
def vendor_of(desc):
    d = desc.split("\n")[0]
    if d.startswith("Payment: "): d = d[9:]
    return re.sub(r"\s+", " ", d.split(" - ")[0].strip().rstrip(",").strip())

opex_vm = collections.defaultdict(lambda: [0.0]*12)
opex_va = collections.defaultdict(collections.Counter)
opex_vd = collections.defaultdict(collections.Counter)
totals = collections.defaultdict(lambda: [0.0]*12)
for acct, m, desc, debit, credit, dept in records:
    c = code(acct)
    if not c or not c[0].isdigit(): continue
    net = debit - credit
    if c in REVENUE_SALES: totals["revenue"][m-1] += -net
    elif c in OTHER_INCOME: totals["other_income"][m-1] += -net
    elif c in COGS: totals["cogs"][m-1] += net
    elif c in PAYROLL_PL: totals["payroll"][m-1] += net
    elif is_opex(c):
        totals["opex"][m-1] += net
        v = vendor_of(desc)
        opex_vm[v][m-1] += net
        opex_va[v][acct] += abs(net)
        opex_vd[v][dept or "(none)"] += abs(net)

# ---------------- 2. contracts + allocations ----------------
contracts = {}
with open(A["contracts"]) as f:
    for row in csv.DictReader(f):
        name = row["Name"].strip()
        if not name: continue
        monthly = money(row["Monthly Cost"]) or money(row["Annual Cost"]) / 12.0
        contracts[name] = dict(monthly=round(monthly,2), unalloc=money(row["Unallocated Amount (monthly)"]),
            unalloc_dept=row["Bill Unallocated Licenses To What Department?"].strip(),
            cost_type=row["Cost Type (drop down)"].strip(), purpose=row["Purpose (short text)"].strip())
fixed_alloc = collections.defaultdict(lambda: collections.defaultdict(float))
seat_alloc = collections.defaultdict(float)
with open(A["allocations"]) as f:
    for row in csv.DictReader(f):
        vend = row["Contracts & Expense"].strip()
        if not vend: continue
        amt = money(row["Allocation Amount"])
        dept, poss = row["Department"].strip(), row["Position(s)"].strip()
        if poss and not dept: seat_alloc[vend] += amt
        elif dept: fixed_alloc[vend][dept] += amt
for name, c in contracts.items():
    if c["unalloc"] and c["unalloc_dept"]:
        fixed_alloc[name][c["unalloc_dept"]] += c["unalloc"]

# ---------------- 3. canonicalize + merge ----------------
EXACT = {"hubspot":"Hubspot","shipedge":"ShipEdge","adjust ppc spend to match month":"Google Advertising",
 "google":"Google Advertising","feb 2026":"Auto Owners Insurance","11 richland llc":"Rent",
 "secure vision solutions":"Install Services Site Surveys","sitetech solutions":"Install Services Site Surveys",
 "truist":"Bank & Credit Card Fees","truist bank":"Bank & Credit Card Fees","amex":"Bank & Credit Card Fees",
 "ipayment":"Bank & Credit Card Fees","chatgpt":"ChatGPT (OpenAI subscriptions)",
 "openai":"OPENAI (Survail LLM)","amazon":"Amazon (supplies)","sales":"Fraud / Disputed Charges",
 "apple":"Fraud / Disputed Charges","capital one":"Capital One Rewards & Credits",
 "sba":"Small Business Administration","aws":"Amazon Web Services",
 "mountain valley spring water":"Mountain Valley Spring Water"}
PREFIX = [("gong","Gong.io Inc"),("zoom","Zoom"),("clickup","ClickUp"),("taxjar","TaxJar"),
 ("dun & bradstreet","Dun & Bradstreet"),("docker","Docker, Inc"),("corporate filings","Corporate Filings LLC"),
 ("adobe","Adobe"),("verizon","Verizon Wireless"),("knack","Knack"),("rippling","Rippling"),
 ("# 015493633","Auto Owners Insurance"),("atlassian","Atlassian"),("dkl investments","DKL Investments, LLC"),
 ("2025-2026 buncombe","Buncombe County Property Tax"),("buncombe county","Buncombe County Property Tax"),
 ("make.com","Make.com"),("digitalocean","DigitalOcean.com"),("logmein","Logmein (Go To Meeting)"),
 ("name-cheap","Name-Cheap.com"),("authorize.net","Authorize.net"),("esignatures","ESignatures.io")]
def canon(v):
    k = v.lower().strip()
    if k in EXACT: return EXACT[k]
    for p, c in PREFIX:
        if k.startswith(p): return c
    for c in contracts:
        if c.lower() == k: return c
    return v

merged = collections.defaultdict(lambda: {"months":[0.0]*12,"accounts":collections.Counter(),"xdepts":collections.Counter()})
for v, months in opex_vm.items():
    cv = canon(v); m = merged[cv]
    for i in range(12): m["months"][i] += months[i]
    tot = abs(sum(months))
    m["accounts"][opex_va[v].most_common(1)[0][0]] += tot
    m["xdepts"][opex_vd[v].most_common(1)[0][0]] += tot

XD2BD = {"Facilities":"Facillities","Administration":"03 - HR & Business Administration Manager",
 "Fulfillment / Warehouse":"Purchasing & Fulfillment","Support":"Technical Support",
 "Installation Services - National":"Installation Services","Installation Services - Asheville":"Installation Services",
 "Installation Services - Triad":"Installation Services","(none)":""}

FULLM = cut.month - 1  # complete months
def classify(months):
    act = [i+1 for i in range(cut.month) if abs(months[i]) > 0.005]
    if not act: return ("INACTIVE", 0.0)
    first, last, n = act[0], act[-1], len(act)
    tot = sum(months[:cut.month]); ytd_mo = tot / ELAPSED
    avg3 = statistics.mean(months[FULLM-3:FULLM]); prior3 = statistics.mean(months[FULLM-6:FULLM-3])
    gaps = [b-a for a, b in zip(act, act[1:])]
    if n <= 2 and (last-first) < 4 and first < FULLM-2: status = "ONE-TIME"
    elif 2 <= n <= 4 and gaps and max(gaps) >= 2 and (last-first) >= 4: status = "PERIODIC"
    elif first >= FULLM-2: status = "NEW"
    elif last <= FULLM-1: status = "DROPPED"
    elif prior3 > 100 and avg3 < 0.6*prior3: status = "DROPPING"
    elif prior3 > 100 and avg3 > 1.4*prior3: status = "GROWING"
    else: status = "STEADY"
    if status == "DROPPED": base = 0.0
    elif status in ("ONE-TIME","PERIODIC"): base = ytd_mo
    elif status == "NEW":
        span = [months[i] for i in range(first-1, FULLM)]
        base = statistics.mean(span) if span else months[cut.month-1]/(cut.day/31.0)
    else: base = avg3
    return (status, round(base, 2))

vend_master = []
for cv, m in sorted(merged.items(), key=lambda kv: -abs(sum(kv[1]["months"][:8]))):
    tot = sum(m["months"][:8])
    if abs(tot) < 100: continue
    status, base = classify(m["months"])
    xdept = XD2BD.get(m["xdepts"].most_common(1)[0][0], m["xdepts"].most_common(1)[0][0])
    fixed = {k: round(v, 2) for k, v in fixed_alloc.get(cv, {}).items()}
    seats = round(seat_alloc.get(cv, 0.0), 2)
    bdept = max(fixed, key=fixed.get) if fixed else ""
    final = bdept or contracts.get(cv, {}).get("unalloc_dept", "") or xdept or "Unassigned"
    vend_master.append(dict(vendor=cv, months=[round(x,2) for x in m["months"][:8]], ytd=round(tot,2),
        account=m["accounts"].most_common(1)[0][0], xero_dept=xdept, budget_dept=bdept,
        status=status, base=base, in_contracts=cv in contracts,
        budget_monthly=contracts.get(cv, {}).get("monthly", 0.0),
        cost_type=contracts.get(cv, {}).get("cost_type", ""), purpose=contracts.get(cv, {}).get("purpose", ""),
        fixed_alloc=fixed, seat_monthly=seats, final_dept=final,
        contradiction=bool(bdept) and bool(xdept) and bdept != xdept))

xnames = {r["vendor"] for r in vend_master}
missing = []
for c, info in contracts.items():
    if c in xnames: continue
    fixed = {k: round(v, 2) for k, v in fixed_alloc.get(c, {}).items()}
    seats = round(seat_alloc.get(c, 0.0), 2)
    if not fixed and seats > 0: continue        # pure per-seat -> lives on STAFFING
    bdept = max(fixed, key=fixed.get) if fixed else (info["unalloc_dept"] or "Unassigned")
    missing.append(dict(vendor=c, budget_monthly=info["monthly"], fixed_alloc=fixed, seat_monthly=seats,
        final_dept=bdept, cost_type=info["cost_type"], purpose=info["purpose"]))

# dept x month actual matrix; small-vendor remainder -> Unassigned so it ties to booked opex
dept_actual = collections.defaultdict(lambda: [0.0]*8)
for r in vend_master:
    for i in range(8): dept_actual[r["final_dept"]][i] += r["months"][i]
for i in range(8):
    resid = totals["opex"][i] - sum(v[i] for v in dept_actual.values())
    dept_actual["Unassigned"][i] += resid
dept_actual = {k: [round(x, 2) for x in v] for k, v in dept_actual.items()}

# ---------------- 4. legacy seeds ----------------
lw = openpyxl.load_workbook(A["legacy"], data_only=True)
st = lw["STAFFING"]
staff = []
for row in st.iter_rows(min_row=3, max_row=40):
    dept, pos = row[0].value, row[1].value
    if not pos: continue
    counts = [row[i].value for i in range(16, min(len(row), 48), 3) if row[i].value is not None]
    staff.append(dict(dept=dept, position=pos, ote=round(float(row[2].value or 0), 2),
        hdv=round(float(row[4].value or 0), 2), lic=round(float(row[5].value or 0), 4),
        tax=float(row[6].value or 0), hc=float(counts[-1] if counts else 0)))
rv = lw["PROJECTED REVENUE CALCULATIONS"]
sales = []
for row in rv.iter_rows(min_row=3, max_row=19):
    pos, name, ramp, start = row[0].value, row[1].value, row[4].value, row[5].value
    if not pos or not ramp: continue
    sales.append(dict(role=pos, name=(name or pos).strip(), ramp=ramp, start_cal=int(4 + (start or 0)),
        ach=float(row[3].value or 1.0)))
ramps = []
for row in lw["RAMP SCHEDULES"].iter_rows(min_row=2, max_row=8):
    if row[0].value:
        ramps.append(dict(name=row[0].value, vals=[float(row[i].value or 0) for i in range(1, 13)]))
lw.close()

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "model_data.json")
json.dump(dict(vend_master=vend_master, missing=missing, dept_actual=dept_actual,
    totals={k: [round(x, 2) for x in v] for k, v in totals.items()},
    staff=staff, sales=sales, ramps=ramps,
    seat_alloc={k: round(v, 2) for k, v in seat_alloc.items()}, cutoff=CUTOFF), open(out, "w"), indent=1)
print("staff hc sample:", [(s["position"][:24], s["hc"]) for s in staff[:5]])
print("rows:", len(vend_master), "missing:", len(missing), "staff:", len(staff), "sales:", len(sales))
print("jan dept sum:", round(sum(v[0] for v in dept_actual.values()),2), "vs opex", totals["opex"][0])
