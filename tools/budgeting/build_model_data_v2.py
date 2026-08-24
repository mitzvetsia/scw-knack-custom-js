#!/usr/bin/env python3
"""Stage 1 (v2): TTM accrual model data for the dashboard.

Sources:
  - Xero "Account Transactions" export (Accrual Basis, Aug 1 2025 - Jul 31 2026):
    authoritative for expenses / COGS / payroll, 12 full months of trend history,
    clean Contact names. Verified identical to the GL Detail on every overlapping
    month+bucket (delta $0.00), so this substitutes cleanly.
  - Xero "General Ledger Detail" export (Jan-Aug 2026): revenue + other income
    (the Account Transactions report filter excluded Revenue accounts) and the
    Departments tracking category per vendor (no dept column in the new report).
  - contractsexpenses.csv + allocations.csv + legacy BUDGET workbook: unchanged.

Output: model_data_v2.json (dashboard input). The v1 model_data.json (workbook) stays.
"""
import sys, json, csv, re, collections, statistics, os
from datetime import datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
A = dict(
    at="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/c384c446-Security_Camera_Warehouse__INC__Account_Transactions_13.xlsx",
    at2="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/bcc7e877-Security_Camera_Warehouse__INC__Account_Transactions_14.xlsx",
    contracts="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/ccd2bd72-contractsexpenses.csv",
    allocations="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/47bec46b-allocations.csv",
    legacy="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/25b0d1bd-BUDGET_TEST_1774983468__Q2Q3_2026_JULY_REWORK.xlsx",
    gl="/root/.claude/uploads/77ac7e98-f266-57ce-b2c5-93304bdfc682/7cf4d3fa-Security_Camera_Warehouse__INC__General_Ledger_Detail_2.xlsx",
    v1=os.path.join(HERE, "model_data.json"),
)
for arg in sys.argv[1:]:
    k, _, v = arg.partition("="); A[k] = v

# TTM month index: 0 = Aug 2025 ... 11 = Jul 2026
TTM_LABELS = ["Aug 25","Sep 25","Oct 25","Nov 25","Dec 25","Jan 26","Feb 26","Mar 26","Apr 26","May 26","Jun 26","Jul 26"]
def ttm_idx(y, m):
    i = (y - 2025) * 12 + m - 8
    return i if 0 <= i <= 11 else None

def money(s):
    if s is None: return 0.0
    s = str(s).replace("$","").replace(",","").strip()
    if not s: return 0.0
    neg = s.startswith("-") or (s.startswith("(") and s.endswith(")"))
    s = s.strip("()-")
    try: v = float(s)
    except ValueError: return 0.0
    return -v if neg else v

# ---------------- Account Transactions (TTM accrual) ----------------
# Two exports are merged. AT#13 (Aug 2025-Jul 2026) is the only source for the 2025 half of
# the TTM window; AT#14 (calendar 2026, pulled 2026-08-24) supersedes it for every 2026 month.
# They disagree on May-Jul 2026 because Xero books DAILY payroll-liability accruals
# ("Payroll Liabilities - Monday/Tuesday/...") that are later reversed in a lump
# ("Reversal of Payroll Liabilities"). AT#13 caught those pairs mid-cycle, understating
# Jun ($233K) and Jul ($286K); the fully-posted figures are $313K and $321K, in line with
# every other month. Always prefer the newest export for a month it covers.
SKIP = ("Total","Opening Balance","Closing Balance","Net movement","No transactions","Date","Account Transactions","Security Camera","For the period","Accrual Basis","Account Type")
def read_at(path, keep):
    """keep(year, month) -> bool. Returns [(account, ttm_idx, contact, debit, credit)]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    out, section = [], None
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        rest = [v for v in row[1:] if v not in (None, "")]
        if isinstance(a, str) and not rest:
            s = a.strip()
            if not any(s.startswith(k) for k in SKIP):
                section = s.lstrip("- ").strip()
            continue
        if isinstance(a, datetime):
            _, contact, desc, ref, gross, debit, credit, runbal, rel = row[:9]
            i = ttm_idx(a.year, a.month)
            if i is None or not keep(a.year, a.month): continue
            out.append((section, i, (contact or "").strip() or (desc or "").strip(),
                        float(debit or 0), float(credit or 0)))
    wb.close()
    return out

records = read_at(A["at"], lambda y, m: y == 2025)          # Aug-Dec 2025 only
if A.get("at2"):
    records += read_at(A["at2"], lambda y, m: y == 2026)    # Jan-Jul 2026 (ttm_idx caps at Jul)
else:
    records += read_at(A["at"], lambda y, m: y == 2026)

def code(acct): return acct.split(" - ")[0].strip() if acct else ""
VARCOMP_RE = re.compile(r"variable comp|commission|bonus|spiff", re.I)
ACCRUAL_RE = re.compile(r"payroll liabilit", re.I)
REVERSAL_RE = re.compile(r"reversal", re.I)
PAYROLL_PL = {"66000","6529","6727","6189"}
COGS = {"5257","5258","5259","5260","5566","5567","5568","5569","5570","5571","6734","6737","6730"}
def is_opex(c):
    return (c.startswith(("6","8")) or c == "208") and c not in PAYROLL_PL and c not in COGS and c != "8311"

opex_vm = collections.defaultdict(lambda: [0.0]*12)   # vendor -> ttm months
opex_va = collections.defaultdict(collections.Counter)
totals_ttm = collections.defaultdict(lambda: [0.0]*12)
cogs_accounts = collections.defaultdict(lambda: [0.0]*12)
for acct, i, contact, debit, credit in records:
    c = code(acct)
    if not c or not c[0].isdigit(): continue
    net = debit - credit
    if c in COGS:
        totals_ttm["cogs"][i] += net
        cogs_accounts[acct][i] += net
    elif c in PAYROLL_PL:
        totals_ttm["payroll"][i] += net
        if c == "66000": totals_ttm["wages"][i] += net
        elif c == "6529": totals_ttm["ptax"][i] += net
        # Variable comp (AE/AM/BDR commissions) is booked INSIDE payroll, so booked months are
        # already total comp cost. Split it out so the dashboard can compare its base-salary
        # model against a like-for-like base figure.
        if VARCOMP_RE.search(contact): totals_ttm["varcomp"][i] += net
        # Payroll-liability accruals post daily and are reversed on the PAY DATE (biweekly),
        # not at month-end, so each calendar month closes with a stub of accrued-but-unreversed
        # days. The residual below is that stub's month-over-month change: it is real accrual
        # accounting, but it makes any SINGLE month a noisy benchmark (Jan -$106K, Jul +$37K,
        # summing to only +$10K over Jan-Jul). Track it so the dashboard can show why.
        if c == "66000":
            if REVERSAL_RE.search(contact): totals_ttm["accrual_rev"][i] += net
            elif ACCRUAL_RE.search(contact): totals_ttm["accrual_acc"][i] += net
    elif is_opex(c):
        totals_ttm["opex"][i] += net
        v = re.sub(r"\s+"," ", contact.split("\n")[0]).strip()
        if v.startswith("Payment: "): v = v[9:]
        opex_vm[v][i] += net
        opex_va[v][acct] += abs(net)

# ---------------- contracts + allocations (as v1) ----------------
contracts = {}
with open(A["contracts"]) as f:
    for row in csv.DictReader(f):
        name = row["Name"].strip()
        if not name: continue
        monthly = money(row["Monthly Cost"]) or money(row["Annual Cost"]) / 12.0
        contracts[name] = dict(monthly=round(monthly,2), unalloc=money(row["Unallocated Amount (monthly)"]),
            unalloc_dept=row["Bill Unallocated Licenses To What Department?"].strip(),
            cost_type=row["Cost Type (drop down)"].strip(), purpose=row["Purpose (short text)"].strip())
fixed_alloc = collections.defaultdict(lambda: collections.defaultdict(float))
seat_alloc = collections.defaultdict(float)
with open(A["allocations"]) as f:
    for row in csv.DictReader(f):
        vend = row["Contracts & Expense"].strip()
        if not vend: continue
        amt = money(row["Allocation Amount"])
        dept, poss = row["Department"].strip(), row["Position(s)"].strip()
        if poss and not dept: seat_alloc[vend] += amt
        elif dept: fixed_alloc[vend][dept] += amt
for name, c in contracts.items():
    if c["unalloc"] and c["unalloc_dept"]:
        fixed_alloc[name][c["unalloc_dept"]] += c["unalloc"]

EXACT = {"hubspot":"Hubspot","shipedge":"ShipEdge","adjust ppc spend to match month":"Google Advertising",
 "google":"Google Advertising","feb 2026":"Auto Owners Insurance","11 richland llc":"Rent",
 "secure vision solutions":"Install Services Site Surveys","sitetech solutions":"Install Services Site Surveys",
 "truist":"Bank & Credit Card Fees","truist bank":"Bank & Credit Card Fees","amex":"Bank & Credit Card Fees",
 "ipayment":"Bank & Credit Card Fees","chatgpt":"ChatGPT (OpenAI subscriptions)",
 "openai":"OPENAI (Survail LLM)","amazon":"Amazon (supplies)","sales":"Fraud / Disputed Charges",
 "apple":"Fraud / Disputed Charges","capital one":"Capital One Rewards & Credits",
 "sba":"Small Business Administration","aws":"Amazon Web Services",
 "mountain valley spring water":"Mountain Valley Spring Water","american express":"Bank & Credit Card Fees"}
PREFIX = [("adjust ppc","Google Advertising"),("gong","Gong.io Inc"),("zoom","Zoom"),("clickup","ClickUp"),("taxjar","TaxJar"),
 ("dun & bradstreet","Dun & Bradstreet"),("docker","Docker, Inc"),("corporate filings","Corporate Filings LLC"),
 ("adobe","Adobe"),("verizon","Verizon Wireless"),("knack","Knack"),("rippling","Rippling"),
 ("# 015493633","Auto Owners Insurance"),("atlassian","Atlassian"),("dkl investments","DKL Investments, LLC"),
 ("2025-2026 buncombe","Buncombe County Property Tax"),("buncombe county","Buncombe County Property Tax"),
 ("make.com","Make.com"),("digitalocean","DigitalOcean.com"),("logmein","Logmein (Go To Meeting)"),
 ("name-cheap","Name-Cheap.com"),("authorize.net","Authorize.net"),("esignatures","ESignatures.io"),
 ("birdeye","BirdEye Inc"),("linkedin","LinkedIn")]
def canon(v):
    k = v.lower().strip()
    if k in EXACT: return EXACT[k]
    for p, c in PREFIX:
        if k.startswith(p): return c
    for c in contracts:
        if c.lower() == k: return c
    return (v[:57].rstrip() + "…") if len(v) > 60 else v

merged = collections.defaultdict(lambda: {"months":[0.0]*12, "accounts":collections.Counter()})
for v, months in opex_vm.items():
    cv = canon(v); m = merged[cv]
    for i in range(12): m["months"][i] += months[i]
    m["accounts"][opex_va[v].most_common(1)[0][0]] += abs(sum(months))

# xero dept per canonical vendor: reuse v1 (parsed from GL Detail's Departments column)
v1 = json.load(open(A["v1"]))
xdept_map = {r["vendor"]: r["xero_dept"] for r in v1["vend_master"]}

def classify(months):
    act = [i for i in range(12) if abs(months[i]) > 0.005]
    if not act: return ("INACTIVE", 0.0)
    first, last, n = act[0], act[-1], len(act)
    tot = sum(months); per_mo = tot / 12.0
    last3 = statistics.mean(months[9:12]); prior3 = statistics.mean(months[6:9])
    gaps = [b - a for a, b in zip(act, act[1:])]
    if n <= 2 and (last - first) < 4 and first < 9: status = "ONE-TIME"
    elif 2 <= n <= 5 and gaps and max(gaps) >= 2 and (last - first) >= 6: status = "PERIODIC"
    elif first >= 9: status = "NEW"
    elif last <= 9: status = "DROPPED"
    elif prior3 > 100 and last3 < 0.6 * prior3: status = "DROPPING"
    elif prior3 > 100 and last3 > 1.4 * prior3: status = "GROWING"
    else: status = "STEADY"
    if status == "DROPPED": base = 0.0
    elif status in ("ONE-TIME","PERIODIC"): base = per_mo
    elif status == "NEW": base = statistics.mean(months[first:12])
    else: base = last3
    return (status, round(base, 2))

vend_master = []
for cv, m in sorted(merged.items(), key=lambda kv: -abs(sum(kv[1]["months"]))):
    tot = sum(m["months"]); t26 = sum(m["months"][5:12])
    if abs(tot) < 150: continue
    status, base = classify(m["months"])
    xdept = xdept_map.get(cv, "")
    fixed = {k: round(v, 2) for k, v in fixed_alloc.get(cv, {}).items()}
    seats = round(seat_alloc.get(cv, 0.0), 2)
    bdept = max(fixed, key=fixed.get) if fixed else ""
    final = bdept or contracts.get(cv, {}).get("unalloc_dept", "") or xdept or "Unassigned"
    vend_master.append(dict(vendor=cv, ttm=[round(x, 2) for x in m["months"]], ttm_total=round(tot, 2),
        y26=[round(x, 2) for x in m["months"][5:12]],
        account=m["accounts"].most_common(1)[0][0], xero_dept=xdept, budget_dept=bdept,
        status=status, base=base, in_contracts=cv in contracts,
        budget_monthly=contracts.get(cv, {}).get("monthly", 0.0),
        purpose=contracts.get(cv, {}).get("purpose", ""),
        fixed_alloc=fixed, seat_monthly=seats, final_dept=final,
        contradiction=bool(bdept) and bool(xdept) and bdept != xdept))

xnames = {r["vendor"] for r in vend_master}
missing = []
for c, info in contracts.items():
    if c in xnames: continue
    fixed = {k: round(v, 2) for k, v in fixed_alloc.get(c, {}).items()}
    seats = round(seat_alloc.get(c, 0.0), 2)
    if not fixed and seats > 0: continue
    bdept = max(fixed, key=fixed.get) if fixed else (info["unalloc_dept"] or "Unassigned")
    missing.append(dict(vendor=c, budget_monthly=info["monthly"], fixed_alloc=fixed, seat_monthly=seats,
        final_dept=bdept, purpose=info["purpose"]))

# actual payroll by department (GL Detail carries the Departments tracking on payroll postings)
glwb = openpyxl.load_workbook(A["gl"], read_only=True, data_only=True)
glws = glwb.worksheets[0]
GLSKIP = ("Total","Opening","Closing","Net movement","No transactions","Date","General Ledger","For the period")
GLD2BD = {"Facilities":"Facillities","Administration":"03 - HR & Business Administration Manager",
 "Fulfillment / Warehouse":"Purchasing & Fulfillment","Support":"Technical Support",
 "Installation Services - National":"Installation Services","Installation Services - Asheville":"Installation Services",
 "Installation Services - Triad":"Installation Services"}
dept_payroll = collections.defaultdict(lambda: [0.0]*7)
sec = None
for row in glws.iter_rows(values_only=True):
    a = row[0]
    rest = [v for v in row[1:] if v not in (None, "")]
    if isinstance(a, str) and not rest and not a.startswith(GLSKIP) and "period" not in a:
        sec = a.strip(); continue
    if isinstance(a, datetime) and a.year == 2026 and a.month <= 7:
        c = code(sec)
        if c not in PAYROLL_PL: continue
        _, source, desc, ref, debit, credit, runbal, dept, proj, relacct = row[:10]
        d = GLD2BD.get((dept or "").strip(), (dept or "").strip()) or "Unassigned"
        dept_payroll[d][a.month - 1] += float(debit or 0) - float(credit or 0)
glwb.close()
# The GL Detail export predates AT#14, so its month totals carry the same mid-cycle
# payroll-accrual understatement. Keep its DEPARTMENT MIX (the only source of dept tags on
# payroll) but scale each month to the authoritative AT#14 total, so the department rollup
# ties to the P&L payroll row. The deltas are company-wide accrual/reversal lumps, so a
# proportional spread is the right allocation.
pay26 = totals_ttm["payroll"][5:12]
pay_scale = []
for i in range(7):
    gl_tot = sum(v[i] for v in dept_payroll.values())
    s = (pay26[i] / gl_tot) if abs(gl_tot) > 1 else 1.0
    pay_scale.append(round(s, 4))
    if abs(gl_tot) > 1:
        for v in dept_payroll.values(): v[i] *= s
dept_payroll = {k: [round(x, 2) for x in v] for k, v in dept_payroll.items()}

# dept x month actuals for Jan-Jul 2026 (modeling year), residual -> Unassigned
dept_actual = collections.defaultdict(lambda: [0.0]*7)
for r in vend_master:
    for i in range(7): dept_actual[r["final_dept"]][i] += r["y26"][i]
opex26 = totals_ttm["opex"][5:12]
for i in range(7):
    dept_actual["Unassigned"][i] += opex26[i] - sum(v[i] for v in dept_actual.values())
dept_actual = {k: [round(x, 2) for x in v] for k, v in dept_actual.items()}

model = dict(
    ttm_labels=TTM_LABELS,
    vend_master=vend_master, missing=missing, dept_actual=dept_actual,
    totals=dict(
        revenue=v1["totals"]["revenue"][:7], other_income=v1["totals"]["other_income"][:7],
        cogs=[round(x,2) for x in totals_ttm["cogs"][5:12]],
        payroll=[round(x,2) for x in totals_ttm["payroll"][5:12]],
        opex=[round(x,2) for x in opex26],
        ttm_opex=[round(x,2) for x in totals_ttm["opex"]],
        ttm_cogs=[round(x,2) for x in totals_ttm["cogs"]],
        ttm_payroll=[round(x,2) for x in totals_ttm["payroll"]],
        varcomp=[round(x,2) for x in totals_ttm["varcomp"][5:12]],
        wages=[round(x,2) for x in totals_ttm["wages"][5:12]],
        accrual_drift=[round(totals_ttm["accrual_acc"][i] + totals_ttm["accrual_rev"][i], 2)
                       for i in range(5, 12)]),
    staff=v1["staff"], sales=v1["sales"], ramps=v1["ramps"],
    employer_tax_rate=round(sum(totals_ttm["ptax"]) / sum(totals_ttm["wages"]), 4),
    cogs_accounts={k: [round(x, 2) for x in v] for k, v in cogs_accounts.items() if abs(sum(v)) > 500},
    provisional_cogs_months=[7],   # user-confirmed 2026-08-24: July COGS entry still not made
    dept_payroll_actual=dept_payroll,
    # August 2026 is deliberately EXCLUDED (actual_months stays 7). AT#14 carries Aug data but
    # the month is incomplete on every line — payroll posts only through Aug 17 ($112K wages vs
    # $280K in Jul), COGS $94K vs $232K, opex $141K vs $206K, revenue $613K vs ~$1.2M.
    built="2026-08-24", actual_months=7)
json.dump(model, open(os.path.join(HERE, "model_data_v2.json"), "w"))
print("vendors:", len(vend_master), "missing:", len(missing))
print("statuses:", dict(collections.Counter(r["status"] for r in vend_master)))
print("conflicts:", sum(1 for r in vend_master if r["contradiction"]))
print("jan-jul opex ties:", [round(sum(v[i] for v in dept_actual.values()) - opex26[i], 2) for i in range(7)])
print("TTM opex by month:", [round(x/1000,1) for x in totals_ttm["opex"]])
print("size KB:", os.path.getsize(os.path.join(HERE, "model_data_v2.json"))//1024)
