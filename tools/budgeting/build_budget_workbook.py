#!/usr/bin/env python3
"""Build SCW_Budget_Projections_2026.xlsx from model_data.json.

model_data.json is derived from:
  - Xero General Ledger Detail export (Jan 1 - Aug 21, 2026)  [actuals, accrual basis]
  - contractsexpenses.csv + allocations.csv (legacy budgeting tool exports)
  - BUDGET_TEST xlsx (legacy tool: staffing roster, sales roster, ramp schedules)

Where a department allocation in the budgeting tool contradicts Xero's
Departments tracking category, the budgeting tool's allocation wins.
"""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
M = json.load(open(os.path.join(HERE, "model_data.json")))

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
DEPTS = ["Sales","Marketing","Software Development","Installation Services","Technical Support",
         "Purchasing & Fulfillment","03 - HR & Business Administration Manager","Facillities",
         "Exec","Product","General Overhead","Unassigned"]
ELAPSED = 7 + 21/31  # months of 2026 elapsed at export date (Aug 21)

# ---- styles ----
F = lambda **kw: Font(name="Arial", **kw)
TITLE   = F(size=14, bold=True, color="1F3864")
NOTE    = F(size=9, italic=True, color="595959")
HDRF    = F(size=9, bold=True, color="FFFFFF")
HDRFILL = PatternFill("solid", fgColor="1F3864")
SECF    = F(size=10, bold=True, color="1F3864")
SECFILL = PatternFill("solid", fgColor="D9E2F3")
INPUT   = F(size=10, color="0000FF")            # blue = edit me
INPUTB  = F(size=10, bold=True, color="0000FF")
DATA    = F(size=10, color="404040")            # dark gray = static Xero/source data
FORM    = F(size=10)                            # black = formula
FORMB   = F(size=10, bold=True)
GREEN   = F(size=10, color="008000")            # link to another sheet
YEL     = PatternFill("solid", fgColor="FFFF00")
LGRAY   = PatternFill("solid", fgColor="F2F2F2")
WARNF   = PatternFill("solid", fgColor="FCE4EC")
OKF     = PatternFill("solid", fgColor="E8F5E9")
EXFILL  = PatternFill("solid", fgColor="EDEDED")
THIN    = Border(bottom=Side(style="thin", color="BFBFBF"))
MONEY   = '$#,##0;($#,##0);"-"'
MONEY2  = '$#,##0.00;($#,##0.00);"-"'
PCT     = '0.0%;(0.0%);"-"'
PCT2    = '0.00%'
STATUS_FILL = {"DROPPED": PatternFill("solid", fgColor="F8CBAD"),
               "DROPPING": PatternFill("solid", fgColor="FCE4D6"),
               "NEW": PatternFill("solid", fgColor="DDEBF7"),
               "GROWING": PatternFill("solid", fgColor="E2EFDA"),
               "ONE-TIME": PatternFill("solid", fgColor="EDEDED"),
               "PERIODIC": PatternFill("solid", fgColor="FFF2CC")}

wb = openpyxl.Workbook()
wb.remove(wb.active)

def put(ws, r, c, v, font=FORM, fill=None, fmt=None, align=None, border=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    if align: cell.alignment = Alignment(horizontal=align)
    if border: cell.border = border
    return cell

def hdr_row(ws, r, cols, start=1):
    for i, t in enumerate(cols):
        put(ws, r, start+i, t, HDRF, HDRFILL, align="center")

# ============================== RAMP SCHEDULES ==============================
rs = wb.create_sheet("RAMP SCHEDULES")
put(rs, 1, 1, "RAMP SCHEDULES — monthly revenue run-rate by tenure month", TITLE)
put(rs, 2, 1, "Blue = editable. Each curve is the expected monthly sales for a rep in their Nth month. "
              "SALES TEAM rows point at a curve by name. Source: legacy budget tool RAMP SCHEDULES tab.", NOTE)
hdr_row(rs, 3, ["CURVE"] + [f"M{i}" for i in range(1,13)] + ["YR1 TOTAL"])
RAMP_FIRST, RAMP_LAST = 4, 4 + len(M["ramps"]) + 3 - 1   # data + 3 blanks
r = RAMP_FIRST
for rp in M["ramps"]:
    put(rs, r, 1, rp["name"], INPUTB)
    for i, v in enumerate(rp["vals"]):
        put(rs, r, 2+i, v, INPUT, fmt=MONEY)
    put(rs, r, 14, f"=SUM(B{r}:M{r})", FORMB, fmt=MONEY)
    r += 1
for _ in range(3):  # blank curve rows
    for c in range(1, 14): put(rs, r, c, None, INPUT, fmt=MONEY if c > 1 else None)
    put(rs, r, 14, f"=SUM(B{r}:M{r})", FORMB, fmt=MONEY)
    r += 1
rs.column_dimensions["A"].width = 16
for c in range(2, 15): rs.column_dimensions[get_column_letter(c)].width = 10
rs.freeze_panes = "B4"

# ============================== XERO ACTUALS ==============================
xa = wb.create_sheet("XERO ACTUALS")
put(xa, 1, 1, "XERO ACTUALS — operating expenses as booked (accrual), Jan 1 – Aug 21 2026", TITLE)
put(xa, 2, 1, "Static data from the Xero General Ledger Detail export. Aug is PARTIAL (through 8/21). "
              "FINAL DEPT: where the budgeting tool's allocation contradicts Xero's department tracking, the budgeting tool wins. "
              "SEAT $/mo = per-seat license portion allocated to positions — that cost is modeled on STAFFING (Licenses/mo), not here. "
              "Payroll, COGS and revenue are summarized in the P&L block at the bottom, not in this vendor list.", NOTE)
XA_HDR = 4
hdr_row(xa, XA_HDR, ["VENDOR / EXPENSE","GL ACCOUNT","XERO DEPT","BUDGET-TOOL DEPT","CONFLICT?","FINAL DEPT"]
        + [m if m != "Aug" else "Aug (partial)" for m in MONTHS[:8]]
        + ["YTD","AVG MAY–JUL","YTD / MO","TREND L3M/P3M","STATUS","SUGGESTED $/MO","BUDGET TOOL $/MO","VAR vs BUDGET","SEAT $/MO (→STAFFING)","PURPOSE / NOTES"])
XA_FIRST = XA_HDR + 1
r = XA_FIRST
for v in M["vend_master"]:
    put(xa, r, 1, v["vendor"], DATA)
    put(xa, r, 2, v["account"], DATA)
    put(xa, r, 3, v["xero_dept"] or "—", DATA)
    put(xa, r, 4, v["budget_dept"] or "—", DATA)
    put(xa, r, 5, "CONFLICT → budget wins" if v["contradiction"] else "", F(size=9, bold=True, color="9C2757"),
        WARNF if v["contradiction"] else None)
    put(xa, r, 6, v["final_dept"], DATA)
    for i in range(8):
        put(xa, r, 7+i, v["months"][i], DATA, fmt=MONEY)
    put(xa, r, 15, f"=SUM(G{r}:N{r})", FORM, fmt=MONEY)
    put(xa, r, 16, f"=AVERAGE(K{r}:M{r})", FORM, fmt=MONEY)
    put(xa, r, 17, f"=O{r}/{ELAPSED:.4f}", FORM, fmt=MONEY)
    put(xa, r, 18, f'=IFERROR(AVERAGE(K{r}:M{r})/AVERAGE(H{r}:J{r}),"")', FORM, fmt='0%;(0%);"-"')
    sc = put(xa, r, 19, v["status"], F(size=9, bold=True), STATUS_FILL.get(v["status"]))
    # suggested $/mo: DROPPED->0; ONE-TIME/PERIODIC->YTD/mo; NEW->build-time avg since first activity; else avg May-Jul
    put(xa, r, 20, f'=IF($S{r}="DROPPED",0,IF(OR($S{r}="ONE-TIME",$S{r}="PERIODIC"),$Q{r},'
                   f'IF($S{r}="NEW",{v["base"]},$P{r})))', FORMB, fmt=MONEY)
    put(xa, r, 21, v["budget_monthly"] if v["in_contracts"] else "", DATA, fmt=MONEY)
    put(xa, r, 22, f'=IF($U{r}="","",$T{r}-$U{r})', FORM, fmt=MONEY)
    put(xa, r, 23, v["seat_monthly"] or "", DATA, fmt=MONEY)
    put(xa, r, 24, v["purpose"], NOTE)
    r += 1
XA_LAST = r - 1
# P&L summary block
r += 2
put(xa, r, 1, "ACTUAL P&L SUMMARY (Xero, monthly)", SECF, SECFILL); PL_TITLE = r
r += 1
hdr_row(xa, r, [""] + MONTHS[:8]); put(xa, r, 9, "Aug (partial)", HDRF, HDRFILL, align="center")
r += 1
T = M["totals"]
def pl_row(label, series, font=DATA, fmt=MONEY):
    global r
    put(xa, r, 1, label, F(size=10, bold=True))
    for i in range(8): put(xa, r, 2+i, round(series[i],2), font, fmt=fmt)
    r += 1
    return r - 1
ROW_REV   = pl_row("Sales revenue", T["revenue"])
ROW_OTH   = pl_row("Other income (rental, interest, rebates)", T["other_income"])
ROW_COGS  = pl_row("COGS (5xxx + install subcontractors 6734)", T["cogs"])
ROW_COGSP = r
put(xa, r, 1, "COGS % of sales revenue", F(size=10, bold=True))
for i in range(8):
    cl = get_column_letter(2+i)
    put(xa, r, 2+i, f"=IFERROR({cl}{ROW_COGS}/{cl}{ROW_REV},0)", FORM, fmt=PCT)
r += 1
pay = [T["payroll"][i] for i in range(8)]
ROW_PAY = pl_row("Payroll as booked (wages+tax+benefits, incl. commissions)", pay)
ROW_DEPT0 = r
for d in DEPTS:
    series = M["dept_actual"].get(d, [0]*8)
    put(xa, r, 1, f"Opex — {d}", DATA)
    for i in range(8): put(xa, r, 2+i, round(series[i],2), DATA, fmt=MONEY)
    r += 1
ROW_OPEXT = r
put(xa, r, 1, "Total operating expenses (non-payroll)", F(size=10, bold=True))
for i in range(8):
    cl = get_column_letter(2+i)
    put(xa, r, 2+i, f"=SUM({cl}{ROW_DEPT0}:{cl}{ROW_OPEXT-1})", FORMB, fmt=MONEY)
r += 1
ROW_NET = r
put(xa, r, 1, "Net profit as booked", F(size=10, bold=True))
for i in range(8):
    cl = get_column_letter(2+i)
    put(xa, r, 2+i, f"={cl}{ROW_REV}+{cl}{ROW_OTH}-{cl}{ROW_COGS}-{cl}{ROW_PAY}-{cl}{ROW_OPEXT}", FORMB, fmt=MONEY)
xa.column_dimensions["A"].width = 34
for c in "BCDEF": xa.column_dimensions[c].width = 15
xa.column_dimensions["E"].width = 20
for c in range(7, 23): xa.column_dimensions[get_column_letter(c)].width = 11
xa.column_dimensions["S"].width = 12
xa.column_dimensions["X"].width = 40
xa.freeze_panes = "B5"

XREF = {"XA_FIRST": XA_FIRST, "XA_LAST": XA_LAST, "ROW_REV": ROW_REV, "ROW_OTH": ROW_OTH,
        "ROW_COGS": ROW_COGS, "ROW_COGSP": ROW_COGSP, "ROW_PAY": ROW_PAY,
        "ROW_DEPT0": ROW_DEPT0, "ROW_OPEXT": ROW_OPEXT}
print("XERO ACTUALS:", XREF)

# ============================== STAFFING ==============================
st = wb.create_sheet("STAFFING")
put(st, 1, 1, "STAFFING — headcount & fully-loaded payroll model (all departments)", TITLE)
put(st, 2, 1, "Blue = editable. Monthly cost = (Salary/12 + $ raise + Salary/12 × (role % + dept %)) × (1 + tax&match %) + Health/mo + Licenses/mo. "
              "Raises apply from RAISE EFF. MONTH onward. Headcount cells: set the number of seats per month — add hires by raising the count from their start month, "
              "or add a whole new role in the blank rows. Sales commissions are NOT here — they are on SALES TEAM. "
              "Roster, salaries, benefits, per-seat license $ and tax % seeded from the legacy budget tool.", NOTE)
ST_MO_ROW = 5
ST_HDR = 6
# month number helper row over headcount (N..Y => cols 14..25) and cost (Z..AK => 26..37)
for i in range(12):
    put(st, ST_MO_ROW, 14+i, i+1, F(size=8, color="A0A0A0"), align="center")
    put(st, ST_MO_ROW, 26+i, i+1, F(size=8, color="A0A0A0"), align="center")
put(st, ST_MO_ROW, 1, "month # →", F(size=8, color="A0A0A0"))
hdr_row(st, ST_HDR, ["DEPARTMENT","POSITION","ANNUAL SALARY / OTE BASE","SALARY /MO","HEALTH (HDV) /MO",
                     "LICENSES /MO (per seat)","TAX + MATCH %","CURRENT MONTHLY COST /HEAD","$ RAISE /MO",
                     "ROLE % RAISE","DEPT % RAISE (auto)","RAISE EFF. MONTH","NEW MONTHLY COST /HEAD"]
        + [f"HC {m}" for m in MONTHS] + [f"COST {m}" for m in MONTHS] + ["FY COST"])
ST_FIRST = ST_HDR + 1
NBLANK_ST = 8
ST_LAST = ST_FIRST + len(M["staff"]) + NBLANK_ST - 1
DEPT_TBL_FIRST = ST_LAST + 10   # computed below precisely
r = ST_FIRST
def staffing_row(r, seed=None):
    dept = seed["dept"] if seed else None
    put(st, r, 1, dept, INPUT)
    put(st, r, 2, seed["position"] if seed else None, INPUT)
    put(st, r, 3, seed["ote"] if seed else None, INPUT, fmt=MONEY)
    put(st, r, 4, f'=IF($C{r}="",0,$C{r}/12)', FORM, fmt=MONEY)
    put(st, r, 5, seed["hdv"] if seed else 523.5, INPUT, fmt=MONEY)
    put(st, r, 6, round(seed["lic"],2) if seed else 75, INPUT, fmt=MONEY2)
    put(st, r, 7, seed["tax"] if seed else 0.10, INPUT, fmt=PCT)
    put(st, r, 8, f'=IF($C{r}="",0,$D{r}*(1+$G{r})+$E{r}+$F{r})', FORM, fmt=MONEY)
    put(st, r, 9, 0, INPUT, fmt=MONEY)
    put(st, r, 10, 0, INPUT, fmt=PCT)
    put(st, r, 11, f'=IFERROR(INDEX($B${{DT0}}:$B${{DT1}},MATCH($A{r},$A${{DT0}}:$A${{DT1}},0)),0)', FORM, fmt=PCT)
    put(st, r, 12, 9, INPUT)
    put(st, r, 13, f'=IF($C{r}="",0,($D{r}+$I{r}+$D{r}*($J{r}+$K{r}))*(1+$G{r})+$E{r}+$F{r})', FORM, fmt=MONEY)
    for i in range(12):
        put(st, r, 14+i, seed["hc"] if seed else None, INPUT, align="center")
        hc_cl = get_column_letter(14+i); mo_cl = get_column_letter(26+i)
        put(st, r, 26+i, f'=IF({hc_cl}{r}="",0,{hc_cl}{r}*IF({mo_cl}${ST_MO_ROW}>=$L{r},$M{r},$H{r}))', FORM, fmt=MONEY)
    put(st, r, 38, f"=SUM(Z{r}:AK{r})", FORMB, fmt=MONEY)
for s in M["staff"]:
    staffing_row(r, s); r += 1
for _ in range(NBLANK_ST):
    staffing_row(r, None); r += 1
ST_LAST = r - 1
ST_TOT = r
put(st, r, 2, "TOTALS", FORMB, SECFILL)
for i in range(12):
    hc_cl = get_column_letter(14+i); mo_cl = get_column_letter(26+i)
    put(st, r, 14+i, f"=SUM({hc_cl}{ST_FIRST}:{hc_cl}{ST_LAST})", FORMB, SECFILL, align="center")
    put(st, r, 26+i, f"=SUM({mo_cl}{ST_FIRST}:{mo_cl}{ST_LAST})", FORMB, SECFILL, fmt=MONEY)
put(st, r, 38, f"=SUM(AL{ST_FIRST}:AL{ST_LAST})", FORMB, SECFILL, fmt=MONEY)
r += 2
# example row (excluded from totals)
put(st, r, 1, "EXAMPLE (not counted): Technical Support", F(size=9, italic=True, color="808080"), EXFILL)
put(st, r, 2, "Support Engineer — new hire Oct", F(size=9, italic=True, color="808080"), EXFILL)
for c, v, fm in [(3, 65000, MONEY),(5, 523.5, MONEY),(6, 75, MONEY2),(7, 0.10, PCT),(9, 0, MONEY),(10, 0, PCT),(12, 9, None)]:
    put(st, r, c, v, F(size=9, italic=True, color="808080"), EXFILL, fmt=fm)
for i in range(12):
    put(st, r, 14+i, 1 if i >= 9 else 0, F(size=9, italic=True, color="808080"), EXFILL, align="center")
put(st, r, 24, "→ headcount 0 until Oct, 1 after = hire lands Oct 1", F(size=9, italic=True, color="808080"))
r += 2
ST_XERO = r
put(st, r, 1, "Xero payroll as booked (wages+tax+benefits — includes sales commissions)", F(size=10, bold=True))
for i in range(8):
    put(st, r, 26+i, round(M["totals"]["payroll"][i],2), DATA, fmt=MONEY)
put(st, r+1, 1, "Model vs booked (Δ; model excludes commissions — expect model ≤ booked in strong sales months)", NOTE)
for i in range(8):
    cl = get_column_letter(26+i)
    put(st, r+1, 26+i, f"={cl}{ST_TOT}-{cl}{ST_XERO}", FORM, fmt=MONEY)
r += 3
DT_HDR = r
put(st, r, 1, "DEPARTMENT-WIDE % RAISE (applies to every role in the dept, from each row's RAISE EFF. MONTH)", SECF, SECFILL)
r += 1
hdr_row(st, r, ["DEPARTMENT","% RAISE"])
r += 1
DT0 = r
for d in DEPTS:
    if d == "Unassigned": continue
    put(st, r, 1, d, FORM)
    put(st, r, 2, 0, INPUT, YEL, fmt=PCT)
    r += 1
DT1 = r - 1
# patch dept-lookup formulas now that DT range known
for rr in range(ST_FIRST, ST_LAST+1):
    st.cell(row=rr, column=11).value = f'=IFERROR(INDEX($B${DT0}:$B${DT1},MATCH($A{rr},$A${DT0}:$A${DT1},0)),0)'
st.column_dimensions["A"].width = 26
st.column_dimensions["B"].width = 34
for c in range(3, 14): st.column_dimensions[get_column_letter(c)].width = 12
for c in range(14, 26): st.column_dimensions[get_column_letter(c)].width = 7
for c in range(26, 39): st.column_dimensions[get_column_letter(c)].width = 11
st.freeze_panes = "C7"
dv_dept = DataValidation(type="list", formula1='"'+",".join(d for d in DEPTS)+'"', allow_blank=True)
st.add_data_validation(dv_dept); dv_dept.add(f"A{ST_FIRST}:A{ST_LAST}")
print("STAFFING: first,last,tot,xero,dt", ST_FIRST, ST_LAST, ST_TOT, ST_XERO, DT0, DT1)

# ============================== SALES TEAM ==============================
sl = wb.create_sheet("SALES TEAM")
put(sl, 1, 1, "SALES TEAM — individual reps, ramps, MODEST vs STRETCH attainment, and comp liability", TITLE)
put(sl, 2, 1, "Blue = editable. Each rep: pick a ramp curve, a start month (may be ≤ 0 if they started before 2026), "
              "then set attainment per scenario — e.g. Zach 80% modest / 100% stretch. Monthly revenue = ramp value for their tenure month × attainment. "
              "COMM % is paid on attained revenue (defaults from the legacy tool: AE 2.5%, AM 1.5%, Sr AE 1.75% — edit freely). "
              "Sales Manager + Director overrides are computed below the table. Base salaries live on STAFFING.", NOTE)
put(sl, 3, 1, "Annual revenue target ($)", F(size=10, bold=True))
put(sl, 3, 2, 12050000, INPUTB, YEL, fmt=MONEY)
put(sl, 3, 4, "SM override % (of all sales)", F(size=10, bold=True))
put(sl, 3, 5, 0.015, INPUTB, YEL, fmt=PCT2)
put(sl, 3, 7, "Director override % (above target)", F(size=10, bold=True))
put(sl, 3, 8, 0.035, INPUTB, YEL, fmt=PCT2)
put(sl, 3, 10, "Monthly target", F(size=10, bold=True))
put(sl, 3, 11, "=$B$3/12", FORM, fmt=MONEY)
SL_MO_ROW = 4
for i in range(12):
    put(sl, SL_MO_ROW, 9+i, i+1, F(size=8, color="A0A0A0"), align="center")   # I..T modest
    put(sl, SL_MO_ROW, 21+i, i+1, F(size=8, color="A0A0A0"), align="center")  # U..AF stretch
SL_HDR = 5
hdr_row(sl, SL_HDR, ["REP / SOURCE","ROLE","RAMP CURVE","START MO (≤0 = pre-2026)","MODEST %","STRETCH %","COMM %","FY RUN-RATE QUOTA"]
        + [f"MOD {m}" for m in MONTHS] + [f"STR {m}" for m in MONTHS]
        + ["FY MODEST","FY STRETCH","FY COMM MOD","FY COMM STR"])
SL_FIRST = SL_HDR + 1
COMM_DEFAULT = {"Account Executive - RAMP": 0.025, "Account Manager - 55K Base": 0.015,
                "Senior Account Executive - Stu Moore": 0.0175, "BEN": 0.0, "MARKETING": 0.0}
r = SL_FIRST
NBLANK_SL = 4
def sales_row(r, seed=None):
    put(sl, r, 1, (seed["name"] if seed else None), INPUT)
    put(sl, r, 2, (seed["role"] if seed else None), INPUT)
    put(sl, r, 3, (seed["ramp"] if seed else None), INPUT)
    put(sl, r, 4, (seed["start_cal"] if seed else None), INPUT, align="center")
    modest = 1.0 if (seed and seed["name"].startswith(("Senior Account Executive","Stu"))) else 0.8
    put(sl, r, 5, (modest if seed else None), INPUT, fmt=PCT)
    put(sl, r, 6, (1.0 if seed else None), INPUT, fmt=PCT)
    put(sl, r, 7, (COMM_DEFAULT.get(seed["role"], 0.025) if seed else None), INPUT, fmt=PCT2)
    put(sl, r, 8, f"=IF($C{r}=\"\",0,12*INDEX('RAMP SCHEDULES'!$B${RAMP_FIRST}:$M${RAMP_LAST},MATCH($C{r},'RAMP SCHEDULES'!$A${RAMP_FIRST}:$A${RAMP_LAST},0),12))", FORM, fmt=MONEY)
    for blk, pcol in ((9, "E"), (21, "F")):
        for i in range(12):
            cl = get_column_letter(blk+i)
            put(sl, r, blk+i,
                f"=IF(OR($C{r}=\"\",{cl}${SL_MO_ROW}<$D{r}),0,"
                f"INDEX('RAMP SCHEDULES'!$B${RAMP_FIRST}:$M${RAMP_LAST},MATCH($C{r},'RAMP SCHEDULES'!$A${RAMP_FIRST}:$A${RAMP_LAST},0),MIN(12,{cl}${SL_MO_ROW}-$D{r}+1))*${pcol}{r}",
                FORM, fmt=MONEY)
    put(sl, r, 33, f"=SUM(I{r}:T{r})", FORMB, fmt=MONEY)
    put(sl, r, 34, f"=SUM(U{r}:AF{r})", FORMB, fmt=MONEY)
    put(sl, r, 35, f"=AG{r}*$G{r}", FORM, fmt=MONEY)
    put(sl, r, 36, f"=AH{r}*$G{r}", FORM, fmt=MONEY)
for s in M["sales"]:
    sales_row(r, s); r += 1
for _ in range(NBLANK_SL):
    sales_row(r, None); r += 1
SL_LAST = r - 1
def team_row(r, label, mk, fmt=MONEY, fill=None):
    put(sl, r, 1, label, FORMB, fill or SECFILL)
    for i in range(12):
        put(sl, r, 9+i, mk(i), FORMB, fill or SECFILL, fmt=fmt)
    put(sl, r, 33, f"=SUM(I{r}:T{r})", FORMB, fill or SECFILL, fmt=fmt)
    return r
RT_MOD = team_row(r, "TEAM REVENUE — MODEST", lambda i: f"=SUM({get_column_letter(9+i)}{SL_FIRST}:{get_column_letter(9+i)}{SL_LAST})"); r += 1
RT_STR = team_row(r, "TEAM REVENUE — STRETCH", lambda i: f"=SUM({get_column_letter(21+i)}{SL_FIRST}:{get_column_letter(21+i)}{SL_LAST})"); r += 1
CM_MOD = team_row(r, "Rep commissions — modest", lambda i: f"=SUMPRODUCT($G${SL_FIRST}:$G${SL_LAST},{get_column_letter(9+i)}${SL_FIRST}:{get_column_letter(9+i)}${SL_LAST})", fill=LGRAY); r += 1
CM_STR = team_row(r, "Rep commissions — stretch", lambda i: f"=SUMPRODUCT($G${SL_FIRST}:$G${SL_LAST},{get_column_letter(21+i)}${SL_FIRST}:{get_column_letter(21+i)}${SL_LAST})", fill=LGRAY); r += 1
SM_MOD = team_row(r, "Sales Manager override — modest", lambda i: f"=$E$3*{get_column_letter(9+i)}${RT_MOD}", fill=LGRAY); r += 1
SM_STR = team_row(r, "Sales Manager override — stretch", lambda i: f"=$E$3*{get_column_letter(9+i)}${RT_STR}", fill=LGRAY); r += 1
DR_MOD = team_row(r, "Director override — modest (above target)", lambda i: f"=$H$3*MAX(0,{get_column_letter(9+i)}${RT_MOD}-$K$3)", fill=LGRAY); r += 1
DR_STR = team_row(r, "Director override — stretch (above target)", lambda i: f"=$H$3*MAX(0,{get_column_letter(9+i)}${RT_STR}-$K$3)", fill=LGRAY); r += 1
CT_MOD = team_row(r, "TOTAL VARIABLE SALES COMP — MODEST", lambda i: f"={get_column_letter(9+i)}{CM_MOD}+{get_column_letter(9+i)}{SM_MOD}+{get_column_letter(9+i)}{DR_MOD}"); r += 1
CT_STR = team_row(r, "TOTAL VARIABLE SALES COMP — STRETCH", lambda i: f"={get_column_letter(9+i)}{CM_STR}+{get_column_letter(9+i)}{SM_STR}+{get_column_letter(9+i)}{DR_STR}"); r += 1
r += 1
put(sl, r, 1, "Note: 'BEN' and 'MARKETING' rows are house/director-sourced revenue lines from the legacy tool (0% rep commission; director is paid via the override above).", NOTE)
sl.column_dimensions["A"].width = 30
sl.column_dimensions["B"].width = 30
for c in range(3, 9): sl.column_dimensions[get_column_letter(c)].width = 12
for c in range(9, 37): sl.column_dimensions[get_column_letter(c)].width = 10
sl.freeze_panes = "C6"
print("SALES TEAM:", SL_FIRST, SL_LAST, "rev", RT_MOD, RT_STR, "comp", CT_MOD, CT_STR)

# ============================== EXPENSES (controls) ==============================
ex = wb.create_sheet("EXPENSES")
put(ex, 1, 1, "EXPENSES — every recurring cost, with ON/OFF, overrides and timing", TITLE)
put(ex, 2, 1, "Blue = editable. PLANNED $/MO picks: OVERRIDE if set → otherwise XERO SUGGESTED or BUDGET TOOL $ per the USE column → × (1 + % ADJ). "
              "Turn an expense off with ON? = N (or end it early with TO MONTH). Add an expense in the blank rows (set DEPT, OVERRIDE $, FROM/TO, ON?=Y). "
              "Departments follow the budgeting tool where it disagrees with Xero. Per-seat license costs (Slack, Google, Bonusly, seat portions of Hubspot/Gong/Zoom/Atlassian/Rippling…) "
              "are NOT rows here — they ride on STAFFING (Licenses/mo × headcount); SEAT $/MO shows what portion was moved there.", NOTE)
EX_MO_ROW = 4
for i in range(12):
    put(ex, EX_MO_ROW, 16+i, i+1, F(size=8, color="A0A0A0"), align="center")
EX_HDR = 5
hdr_row(ex, EX_HDR, ["EXPENSE / VENDOR","DEPARTMENT","SOURCE","STATUS (Xero trend)","SHARE OF VENDOR","SEAT $/MO (→STAFFING)",
                     "XERO SUGGESTED $/MO","BUDGET TOOL $/MO","USE (XERO/BUDGET)","OVERRIDE $/MO","% ADJ","FROM MO","TO MO","ON?","PLANNED $/MO"]
        + MONTHS + ["FY TOTAL"])
EX_FIRST = EX_HDR + 1
dorder = {d: i for i, d in enumerate(DEPTS)}
exp_rows = []
for v in M["vend_master"]:
    fixed = v["fixed_alloc"]
    seat = v["seat_monthly"]
    if len(fixed) > 1:
        tot = sum(fixed.values())
        for d, amt in sorted(fixed.items(), key=lambda kv: -kv[1]):
            share = amt/tot if tot else 0
            exp_rows.append({"vendor": v["vendor"], "dept": d, "source": "XERO+BUDGET", "status": v["status"],
                             "share": round(share,4), "seat": seat, "budget": round(amt,2), "xero": True,
                             "on": "N" if v["status"] in ("DROPPED","ONE-TIME") else "Y",
                             "base_est": max(0, (v["base"]-seat))*share if seat>0 else v["base"]*share})
    else:
        d = v["final_dept"]
        budget = (list(fixed.values())[0] if fixed else (v["budget_monthly"] if v["in_contracts"] else ""))
        exp_rows.append({"vendor": v["vendor"], "dept": d, "source": "XERO+BUDGET" if v["in_contracts"] else "XERO ONLY",
                         "status": v["status"], "share": 1, "seat": seat, "budget": budget, "xero": True,
                         "on": "N" if v["status"] in ("DROPPED","ONE-TIME") else "Y",
                         "base_est": max(0, v["base"]-seat) if seat>0 else v["base"]})
for mrow in M["missing"]:
    fixed = mrow["fixed_alloc"]
    if len(fixed) > 1:
        for d, amt in sorted(fixed.items(), key=lambda kv: -kv[1]):
            exp_rows.append({"vendor": mrow["vendor"], "dept": d, "source": "BUDGET ONLY", "status": "NOT IN XERO 2026",
                             "share": "", "seat": mrow["seat_monthly"], "budget": round(amt,2), "xero": False, "on": "N",
                             "base_est": amt})
    else:
        d = mrow["final_dept"]
        exp_rows.append({"vendor": mrow["vendor"], "dept": d, "source": "BUDGET ONLY", "status": "NOT IN XERO 2026",
                         "share": "", "seat": mrow["seat_monthly"], "budget": round(mrow["budget_monthly"],2),
                         "xero": False, "on": "N", "base_est": mrow["budget_monthly"]})
exp_rows.sort(key=lambda e: (dorder.get(e["dept"], 99), -(abs(e["base_est"]) if isinstance(e["base_est"], (int,float)) else 0)))
NBLANK_EX = 10
r = EX_FIRST
def expense_row(r, e=None):
    put(ex, r, 1, e["vendor"] if e else None, INPUT if not e else (DATA if e["xero"] else FORM))
    put(ex, r, 2, e["dept"] if e else None, INPUT)
    put(ex, r, 3, e["source"] if e else "MANUAL", F(size=9, color="808080"))
    stat = e["status"] if e else ""
    put(ex, r, 4, stat, F(size=9), STATUS_FILL.get(stat))
    put(ex, r, 5, e["share"] if e else "", DATA, fmt="0%")
    put(ex, r, 6, (e["seat"] or "") if e else "", DATA, fmt=MONEY)
    if e and e["xero"]:
        put(ex, r, 7, f"=IFERROR(IF($F{r}>0,MAX(0,INDEX('XERO ACTUALS'!$T${XA_FIRST}:$T${XA_LAST},MATCH($A{r},'XERO ACTUALS'!$A${XA_FIRST}:$A${XA_LAST},0))-$F{r}),"
                      f"INDEX('XERO ACTUALS'!$T${XA_FIRST}:$T${XA_LAST},MATCH($A{r},'XERO ACTUALS'!$A${XA_FIRST}:$A${XA_LAST},0)))*IF($E{r}=\"\",1,$E{r}),\"\")", GREEN, fmt=MONEY)
    else:
        put(ex, r, 7, "", GREEN, fmt=MONEY)
    put(ex, r, 8, (e["budget"] if e else None), DATA if e else INPUT, fmt=MONEY)
    put(ex, r, 9, ("XERO" if (e and e["xero"]) else "BUDGET") if e else None, INPUT, align="center")
    put(ex, r, 10, None, INPUT, fmt=MONEY)
    put(ex, r, 11, 0 if e else None, INPUT, fmt=PCT)
    put(ex, r, 12, 1 if e else None, INPUT, align="center")
    put(ex, r, 13, 12 if e else None, INPUT, align="center")
    put(ex, r, 14, (e["on"] if e else None), INPUTB, align="center")
    put(ex, r, 15, f"=IF($N{r}=\"N\",0,IF($A{r}=\"\",0,(IF($J{r}<>\"\",$J{r},IF($I{r}=\"BUDGET\",IF($H{r}=\"\",0,$H{r}),IF($G{r}=\"\",0,$G{r}))))*(1+IF($K{r}=\"\",0,$K{r}))))", FORMB, fmt=MONEY)
    for i in range(12):
        cl = get_column_letter(16+i)
        put(ex, r, 16+i, f"=IF(AND({cl}${EX_MO_ROW}>=IF($L{r}=\"\",1,$L{r}),{cl}${EX_MO_ROW}<=IF($M{r}=\"\",12,$M{r})),$O{r},0)", FORM, fmt=MONEY)
    put(ex, r, 28, f"=SUM(P{r}:AA{r})", FORMB, fmt=MONEY)
for e in exp_rows:
    expense_row(r, e); r += 1
for _ in range(NBLANK_EX):
    expense_row(r, None); r += 1
EX_LAST = r - 1
EX_TOT = r
put(ex, r, 1, "TOTAL PLANNED OPEX", FORMB, SECFILL)
put(ex, r, 15, f"=SUM(O{EX_FIRST}:O{EX_LAST})", FORMB, SECFILL, fmt=MONEY)
for i in range(12):
    cl = get_column_letter(16+i)
    put(ex, r, 16+i, f"=SUM({cl}{EX_FIRST}:{cl}{EX_LAST})", FORMB, SECFILL, fmt=MONEY)
put(ex, r, 28, f"=SUM(AB{EX_FIRST}:AB{EX_LAST})", FORMB, SECFILL, fmt=MONEY)
r += 2
put(ex, r, 1, "EXAMPLE (not counted): New CRM add-on — dept Sales, OVERRIDE 500, FROM 9, TO 12, ON?=Y → $500/mo Sep–Dec", F(size=9, italic=True, color="808080"), EXFILL)
r += 2
EXD_HDR = r
put(ex, r, 1, "PLANNED OPEX BY DEPARTMENT (feeds PROJECTIONS)", SECF, SECFILL)
r += 1
hdr_row(ex, r, ["DEPARTMENT"] + MONTHS + ["FY"])
r += 1
EXD0 = r
for d in DEPTS:
    put(ex, r, 1, d, FORM)
    for i in range(12):
        cl = get_column_letter(16+i); dcl = get_column_letter(2+i)
        put(ex, r, 2+i, f"=SUMIFS({cl}${EX_FIRST}:{cl}${EX_LAST},$B${EX_FIRST}:$B${EX_LAST},$A{r})", FORM, fmt=MONEY)
    put(ex, r, 14, f"=SUM(B{r}:M{r})", FORMB, fmt=MONEY)
    r += 1
EXD1 = r - 1
put(ex, r, 1, "TOTAL", FORMB, SECFILL)
for i in range(12):
    cl = get_column_letter(2+i)
    put(ex, r, 2+i, f"=SUM({cl}{EXD0}:{cl}{EXD1})", FORMB, SECFILL, fmt=MONEY)
put(ex, r, 14, f"=SUM(N{EXD0}:N{EXD1})", FORMB, SECFILL, fmt=MONEY)
ex.column_dimensions["A"].width = 34
ex.column_dimensions["B"].width = 26
for c in range(3, 16): ex.column_dimensions[get_column_letter(c)].width = 12
ex.column_dimensions["D"].width = 15
for c in range(16, 29): ex.column_dimensions[get_column_letter(c)].width = 10
ex.freeze_panes = "C6"
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ex.add_data_validation(dv_yn); dv_yn.add(f"N{EX_FIRST}:N{EX_LAST}")
dv_use = DataValidation(type="list", formula1='"XERO,BUDGET"', allow_blank=True)
ex.add_data_validation(dv_use); dv_use.add(f"I{EX_FIRST}:I{EX_LAST}")
dv_dept2 = DataValidation(type="list", formula1='"'+",".join(DEPTS)+'"', allow_blank=True)
ex.add_data_validation(dv_dept2); dv_dept2.add(f"B{EX_FIRST}:B{EX_LAST}")
print("EXPENSES:", EX_FIRST, EX_LAST, "dept block", EXD0, EXD1)

# ============================== PROJECTIONS ==============================
pj = wb.create_sheet("PROJECTIONS")
put(pj, 1, 1, "PROJECTIONS — 2026 P&L: Xero actuals through the cutoff month, model after", TITLE)
put(pj, 2, 1, "For leadership review. Columns up to the cutoff show Xero as booked; later columns are driven by STAFFING (payroll), "
              "SALES TEAM (revenue + variable comp, modest vs stretch), and EXPENSES (planned opex). Both scenarios display simultaneously. "
              "Yellow cells are the master levers.", NOTE)
put(pj, 3, 1, "Actuals through month #", F(size=10, bold=True)); put(pj, 3, 2, 7, INPUTB, YEL, align="center")
put(pj, 3, 3, "COGS % fwd", F(size=10, bold=True))
cogs_ytd = sum(M["totals"]["cogs"][:7]) / sum(M["totals"]["revenue"][:7])
put(pj, 3, 4, round(cogs_ytd, 3), INPUTB, YEL, fmt=PCT)
put(pj, 3, 5, "Other income $/mo fwd", F(size=10, bold=True)); put(pj, 3, 6, 3000, INPUTB, YEL, fmt=MONEY)
put(pj, 3, 7, "Labor hedge % fwd", F(size=10, bold=True)); put(pj, 3, 8, 0, INPUTB, YEL, fmt=PCT)
put(pj, 3, 9, "Opex hedge % fwd", F(size=10, bold=True)); put(pj, 3, 10, 0, INPUTB, YEL, fmt=PCT)
put(pj, 3, 11, "(Aug is partial in Xero — leave cutoff at 7 unless data refreshed)", NOTE)
for i in range(12):
    put(pj, 4, 2+i, i+1, F(size=8, color="A0A0A0"), align="center")
hdr_row(pj, 5, ["2026"] + MONTHS + ["FY 2026"])
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, 6, 2+i, f'=IF({cl}$4<=$B$3,"ACTUAL","PROJ")', F(size=8, bold=True, color="808080"), align="center")

def colmap(i, first_col):  # projection col i (0=Jan) -> letter in target block
    return get_column_letter(first_col + i)
def proj_row(r, label, actual_ref, proj_ref, fmt=MONEY, bold=False, fill=None, sumfy=True):
    put(pj, r, 1, label, FORMB if bold else FORM, fill)
    for i in range(12):
        cl = get_column_letter(2+i)
        act = actual_ref(i) if callable(actual_ref) else actual_ref
        prj = proj_ref(i)
        put(pj, r, 2+i, f"=IF({cl}$4<=$B$3,{act if act else 0},{prj})", FORMB if bold else FORM, fill, fmt=fmt)
    if sumfy: put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, fill, fmt=fmt)
    return r

r = 8
put(pj, r, 1, "REVENUE", SECF, SECFILL); r += 1
P_REV_MOD = proj_row(r, "Sales revenue — MODEST",
    lambda i: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_REV}", lambda i: f"'SALES TEAM'!{colmap(i,9)}${RT_MOD}"); r += 1
P_REV_STR = proj_row(r, "Sales revenue — STRETCH",
    lambda i: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_REV}", lambda i: f"'SALES TEAM'!{colmap(i,9)}${RT_STR}"); r += 1
P_OTH = proj_row(r, "Other income (rental, interest, rebates)",
    lambda i: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_OTH}", lambda i: "$F$3"); r += 1
P_TREV_MOD = r
put(pj, r, 1, "TOTAL REVENUE — MODEST", FORMB, LGRAY)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_REV_MOD}+{cl}{P_OTH}", FORMB, LGRAY, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, LGRAY, fmt=MONEY); r += 1
P_TREV_STR = r
put(pj, r, 1, "TOTAL REVENUE — STRETCH", FORMB, LGRAY)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_REV_STR}+{cl}{P_OTH}", FORMB, LGRAY, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, LGRAY, fmt=MONEY); r += 2

put(pj, r, 1, "COST OF GOODS & INSTALL DELIVERY", SECF, SECFILL); r += 1
P_COGSP = proj_row(r, "COGS % of sales revenue",
    lambda i: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_COGSP}", lambda i: "$D$3", fmt=PCT, sumfy=False)
put(pj, r, 14, f"=IFERROR(N{r+1}/N{P_REV_MOD},0)", FORM, fmt=PCT); r += 1
P_COGS_MOD = r
put(pj, r, 1, "COGS — MODEST", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=IF({cl}$4<=$B$3,'XERO ACTUALS'!{colmap(i,2)}${ROW_COGS},{cl}{P_REV_MOD}*{cl}{P_COGSP})", FORM, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, fmt=MONEY); r += 1
P_COGS_STR = r
put(pj, r, 1, "COGS — STRETCH", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=IF({cl}$4<=$B$3,'XERO ACTUALS'!{colmap(i,2)}${ROW_COGS},{cl}{P_REV_STR}*{cl}{P_COGSP})", FORM, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, fmt=MONEY); r += 1
P_GP_MOD = r
put(pj, r, 1, "GROSS PROFIT — MODEST", FORMB, LGRAY)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_TREV_MOD}-{cl}{P_COGS_MOD}", FORMB, LGRAY, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, LGRAY, fmt=MONEY); r += 1
P_GP_STR = r
put(pj, r, 1, "GROSS PROFIT — STRETCH", FORMB, LGRAY)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_TREV_STR}-{cl}{P_COGS_STR}", FORMB, LGRAY, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, LGRAY, fmt=MONEY); r += 2

put(pj, r, 1, "PAYROLL & SALES COMP", SECF, SECFILL); r += 1
P_PAY = proj_row(r, "Payroll — staffing model (base + benefits + licenses)",
    lambda i: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_PAY}", lambda i: f"STAFFING!{colmap(i,26)}${ST_TOT}*(1+$H$3)"); r += 1
P_COMP_MOD = proj_row(r, "Variable sales comp — MODEST (in booked payroll for actual months)",
    None, lambda i: f"'SALES TEAM'!{colmap(i,9)}${CT_MOD}"); r += 1
P_COMP_STR = proj_row(r, "Variable sales comp — STRETCH (in booked payroll for actual months)",
    None, lambda i: f"'SALES TEAM'!{colmap(i,9)}${CT_STR}"); r += 2

put(pj, r, 1, "OPERATING EXPENSES (dept view — budgeting-tool allocations win)", SECF, SECFILL); r += 1
P_DEPT0 = r
for di, d in enumerate(DEPTS):
    proj_row(r, d, lambda i, di=di: f"'XERO ACTUALS'!{colmap(i,2)}${ROW_DEPT0+di}",
             lambda i, di=di: f"EXPENSES!{colmap(i,2)}${EXD0+di}*(1+$J$3)")
    r += 1
P_OPEXT = r
put(pj, r, 1, "TOTAL OPERATING EXPENSES", FORMB, LGRAY)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=SUM({cl}{P_DEPT0}:{cl}{P_OPEXT-1})", FORMB, LGRAY, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, LGRAY, fmt=MONEY); r += 2

put(pj, r, 1, "BOTTOM LINE", SECF, SECFILL); r += 1
P_NET_MOD = r
put(pj, r, 1, "NET PROFIT — MODEST", FORMB, OKF)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_GP_MOD}-{cl}{P_PAY}-{cl}{P_COMP_MOD}-{cl}{P_OPEXT}", FORMB, OKF, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, OKF, fmt=MONEY); r += 1
P_NET_STR = r
put(pj, r, 1, "NET PROFIT — STRETCH", FORMB, OKF)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"={cl}{P_GP_STR}-{cl}{P_PAY}-{cl}{P_COMP_STR}-{cl}{P_OPEXT}", FORMB, OKF, fmt=MONEY)
put(pj, r, 14, f"=SUM(B{r}:M{r})", FORMB, OKF, fmt=MONEY); r += 1
put(pj, r, 1, "Net margin — modest", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=IFERROR({cl}{P_NET_MOD}/{cl}{P_TREV_MOD},0)", FORM, fmt=PCT)
put(pj, r, 14, f"=IFERROR(N{P_NET_MOD}/N{P_TREV_MOD},0)", FORM, fmt=PCT); r += 1
put(pj, r, 1, "Net margin — stretch", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=IFERROR({cl}{P_NET_STR}/{cl}{P_TREV_STR},0)", FORM, fmt=PCT)
put(pj, r, 14, f"=IFERROR(N{P_NET_STR}/N{P_TREV_STR},0)", FORM, fmt=PCT); r += 1
put(pj, r, 1, "Break-even sales revenue (payroll + opex, excl. variable comp)", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=IFERROR(({cl}{P_PAY}+{cl}{P_OPEXT}-{cl}{P_OTH})/(1-{cl}{P_COGSP}),0)", FORM, fmt=MONEY)
r += 1
put(pj, r, 1, "Cumulative net — modest", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=SUM($B${P_NET_MOD}:{cl}{P_NET_MOD})", FORM, fmt=MONEY)
r += 1
put(pj, r, 1, "Cumulative net — stretch", FORM)
for i in range(12):
    cl = get_column_letter(2+i)
    put(pj, r, 2+i, f"=SUM($B${P_NET_STR}:{cl}{P_NET_STR})", FORM, fmt=MONEY)
r += 1
pj.column_dimensions["A"].width = 46
for c in range(2, 15): pj.column_dimensions[get_column_letter(c)].width = 11
pj.freeze_panes = "B6"

# ============================== README ==============================
rd = wb.create_sheet("README")
lines = [
 ("SCW BUDGETING & PROJECTIONS TOOL — 2026", TITLE),
 ("Built 2026-08-21 from: Xero General Ledger Detail (Jan 1 – Aug 21 2026, accrual), the legacy budgeting tool exports "
  "(contractsexpenses.csv + allocations.csv), and the legacy BUDGET workbook (staffing roster, sales roster, ramp schedules).", NOTE),
 ("", None),
 ("HOW THE TABS FIT TOGETHER", SECF),
 ("PROJECTIONS — the leadership view. Actual months show Xero as booked; projected months are driven by the three control tabs. "
  "Modest and stretch scenarios display side by side.", FORM),
 ("SALES TEAM — one row per rep. Set MODEST % and STRETCH % attainment per person (e.g. Zach 80% / Stu 100%), ramp curve and start month. "
  "Rep commissions + Sales Manager + Director overrides = the comp liability that flows to PROJECTIONS.", FORM),
 ("STAFFING — headcount × fully-loaded monthly cost for every role. Model hires/exits by editing monthly headcount or adding rows; "
  "raise levers: per-role $ or %, or department-wide % (table at the bottom), from a chosen month.", FORM),
 ("EXPENSES — every recurring cost with ON/OFF, XERO/BUDGET source pick, $ override, % adjust, and FROM/TO months. "
  "Blank rows at the bottom to add new expenses.", FORM),
 ("XERO ACTUALS — expenses as booked, by vendor by month, with trend status (STEADY / GROWING / DROPPING / DROPPED / NEW / ONE-TIME / PERIODIC) "
  "and a suggested forward $/mo. Bottom block = actual P&L summary that PROJECTIONS reads.", FORM),
 ("RAMP SCHEDULES — editable revenue ramp curves referenced by SALES TEAM.", FORM),
 ("", None),
 ("COLOR LEGEND", SECF),
 ("BLUE text = yours to edit.  YELLOW fill = master levers.  Black = formulas (leave).  GREEN = pulls from another tab.  "
  "Dark gray = static data imported from Xero/the budget tool (refresh by re-running the build script).", FORM),
 ("", None),
 ("RULES & ASSUMPTIONS BAKED IN", SECF),
 ("1. Where the budgeting tool allocates a vendor to a different department than Xero's tracking category, the budgeting tool wins "
  "(20 conflicts found — see CONFLICT column on XERO ACTUALS).", FORM),
 ("2. Per-seat licenses (Slack, Google Workspace, Bonusly, and the seat portions of Hubspot/Gong/Zoom/Atlassian/Rippling/ClickUp) are modeled as "
  "LICENSES/MO on STAFFING so they scale with headcount — they are deliberately absent from the EXPENSES rows (SEAT $/MO column shows the amount moved).", FORM),
 ("3. Payroll for actual months = Xero as booked (includes commissions). Projected months = STAFFING model + SALES TEAM variable comp, "
  "so commission rows show 0 in actual months to avoid double counting.", FORM),
 ("4. COGS (5xxx accounts + 6734 install subcontractors) is modeled as % of sales revenue; the forward % lever defaults to the 2026 YTD actual. "
  "Site surveys (6738) are treated as a Sales operating expense per the budgeting tool's allocation.", FORM),
 ("5. Commission defaults come from the legacy tool: AE 2.5%, AM 1.5%, Senior AE 1.75%, Sales Manager override 1.5% of all sales, "
  "Director 3.5% of revenue above target ($12.05M default = legacy H1 6.35M + H2 5.7M). All editable in blue.", FORM),
 ("6. Tax + matching % per role is carried from the legacy tool (mostly 1%). If you want a fuller employer-burden rate, edit column G on STAFFING.", FORM),
 ("7. August Xero data is partial (through Aug 21) — the actuals cutoff on PROJECTIONS defaults to month 7 (July).", FORM),
 ("8. 'Suggested $/mo' per vendor: DROPPED → $0; ONE-TIME & PERIODIC → YTD ÷ 7.68 months; NEW → average since first activity; "
  "otherwise average of May–Jul. Override any of it on EXPENSES.", FORM),
 ("", None),
 ("REFRESHING THE DATA", SECF),
 ("Re-export the Xero GL Detail + the two CSVs, re-run the aggregation + tools/budgeting/build_budget_workbook.py in the repo, "
  "and a fresh workbook is produced. Your edits live in this file, so copy your blue cells forward (or ask Claude to merge).", FORM),
]
r = 1
for text, fnt in lines:
    if text: put(rd, r, 1, text, fnt or FORM)
    r += 1
rd.column_dimensions["A"].width = 150
for row in rd.iter_rows(min_row=1, max_row=r):
    for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")

# order tabs + colors
order = ["README","PROJECTIONS","SALES TEAM","STAFFING","EXPENSES","XERO ACTUALS","RAMP SCHEDULES"]
wb._sheets = [wb[n] for n in order]
for n, col in [("README","808080"),("PROJECTIONS","1F3864"),("SALES TEAM","2E75B6"),("STAFFING","2E75B6"),
               ("EXPENSES","2E75B6"),("XERO ACTUALS","548235"),("RAMP SCHEDULES","BF8F00")]:
    wb[n].sheet_properties.tabColor = col

out = os.path.join(HERE, "SCW_Budget_Projections_2026.xlsx")
wb.save(out)
print("saved", out)
